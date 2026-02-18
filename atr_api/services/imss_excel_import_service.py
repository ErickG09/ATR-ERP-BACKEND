# atr_api/services/imss_excel_import_service.py

from __future__ import annotations

from typing import Any, Dict, List
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from atr_api.extensions import db
from atr_api.models.operator import Operator
from atr_api.models.operator_imss import OperatorIMSS


MONTH_MAP = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def normalize_name(name: str) -> str:
    return " ".join(str(name or "").strip().upper().split())


def detect_header_row(ws) -> int:
    """
    Detecta la fila donde esté la columna 'Nombre'
    """
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if str(cell.value or "").strip().lower() == "nombre":
                return cell.row
    raise ValueError("No se encontró columna 'Nombre' en el Excel.")


def parse_decimal(value: Any) -> Decimal:
    """
    Convierte valores de Excel a Decimal de forma robusta.
    Acepta float, int, Decimal o string con separadores.
    """
    if value is None:
        raise ValueError("Celda vacía")

    try:
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))

        s = str(value).strip()

        # Normalizar separadores: quitar miles
        s = s.replace(",", "")
        return Decimal(s)
    except (InvalidOperation, ValueError, TypeError):
        raise ValueError(f"Valor inválido para monto IMSS: {value}")


def import_imss_from_excel(
    *,
    client_id: int,
    file_storage,
    year: int,
    dry_run: bool = True,
) -> Dict[str, Any]:

    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active

    header_row = detect_header_row(ws)

    headers: Dict[int, str] = {}
    for col in ws[header_row]:
        headers[col.column] = str(col.value or "").strip().lower()

    # Detectar columnas de meses
    month_columns: Dict[int, int] = {}
    name_col_index = None

    for col_idx, header in headers.items():
        if header == "nombre":
            name_col_index = col_idx
        elif header in MONTH_MAP:
            month_columns[col_idx] = MONTH_MAP[header]

    if not name_col_index:
        raise ValueError("No se encontró columna 'Nombre' válida.")
    if not month_columns:
        raise ValueError("No se encontraron columnas de meses (Enero..Diciembre).")

    operators = (
        db.session.query(Operator)
        .filter(Operator.client_id == client_id)
        .all()
    )

    operator_index = {
        normalize_name(op.nombre): op
        for op in operators
    }

    total_rows = 0
    created = 0
    updated = 0
    skipped_empty = 0
    not_found: List[str] = []
    invalid_values: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=header_row + 1):
        name_cell = row[name_col_index - 1].value
        if not name_cell:
            continue

        total_rows += 1

        normalized = normalize_name(name_cell)
        operator = operator_index.get(normalized)

        if not operator:
            not_found.append(normalized)
            continue

        for col_idx, month in month_columns.items():
            cell = row[col_idx - 1]

            if cell.value in (None, ""):
                skipped_empty += 1
                continue

            try:
                monto = parse_decimal(cell.value)
            except ValueError:
                invalid_values.append({
                    "operator": normalized,
                    "month": month,
                    "raw_value": cell.value,
                })
                continue

            existing = (
                db.session.query(OperatorIMSS)
                .filter_by(
                    client_id=client_id,
                    operator_id=operator.id,
                    year=year,
                    month=month,
                )
                .first()
            )

            if existing:
                if Decimal(existing.monto) != monto:
                    existing.monto = monto
                    updated += 1
            else:
                new_row = OperatorIMSS(
                    client_id=client_id,
                    operator_id=operator.id,
                    year=year,
                    month=month,
                    monto=monto,
                )
                db.session.add(new_row)
                created += 1

    if not dry_run:
        db.session.commit()
    else:
        db.session.rollback()

    return {
        "dry_run": dry_run,
        "year": year,
        "total_rows": total_rows,
        "created": created,
        "updated": updated,
        "skipped_empty_cells": skipped_empty,
        "not_found_count": len(not_found),
        "not_found": not_found,
        "invalid_values_count": len(invalid_values),
        "invalid_values": invalid_values[:50],  # evita respuestas enormes
    }

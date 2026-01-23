from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from werkzeug.datastructures import FileStorage
from sqlalchemy.exc import IntegrityError

from atr_api.extensions import db
from atr_api.errors import ApiError
from atr_api.models import Car
from atr_api.schemas.car import sanitize_car_payload, calc_rendimiento_promedio

# lector xlsx
from openpyxl import load_workbook


@dataclass
class RowError:
    row_number: int
    message: str
    data: Dict[str, Any]


# ----------------------------
# Helpers de parsing/normalización
# ----------------------------
def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _norm_header(h: Any) -> str:
    """
    Normaliza headers: "Km Acum" -> "kmacum"
    """
    import re
    from unicodedata import normalize

    s = _norm(h).lower()
    s = normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def _to_bool(v: Any) -> Optional[bool]:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("1", "true", "t", "si", "sí", "s", "y", "yes"):
        return True
    if s in ("0", "false", "f", "no", "n"):
        return False
    return None


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except Exception:
        return None


def _to_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        # excel puede traer 12.0
        return int(float(str(v).strip()))
    except Exception:
        return None


def _read_xlsx_rows(file_storage: FileStorage) -> List[Dict[str, Any]]:
    """
    Lee el primer sheet:
      - primera fila = headers
      - filas siguientes = datos
    Regresa [{_row_number, payload}]
    """
    try:
        wb = load_workbook(file_storage.stream, data_only=True)
    except Exception as e:
        raise ApiError(f"No se pudo leer el Excel: {e}", status_code=400)

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    headers = list(headers or [])
    if not any(_norm(h) for h in headers):
        raise ApiError("El Excel no tiene encabezados (fila 1).", status_code=400)

    # mapping de headers comunes -> keys del payload (car schema)
    # (si tu machote ya trae keys exactas, también funciona: las dejamos pasar)
    header_map = {
        "codigo": "codigo",
        "unidad": "codigo",
        "carro": "codigo",
        "operador": "operador",

        "kmacum": "km_acum",
        "kmacumulados": "km_acum",
        "kms": "km_acum",

        "ltdiesac": "lt_dies_ac",
        "litrosdiesel": "lt_dies_ac",
        "litros": "lt_dies_ac",

        "activo": "activo",

        # flags típicos que has usado en ATR
        "mexico": "mexico",
        "expver": "exp_ver",
        "exportacionveracruz": "exp_ver",
        "importado": "importado",
    }

    norm_headers: List[str] = []
    for h in headers:
        nh = _norm_header(h)
        norm_headers.append(nh)

    out: List[Dict[str, Any]] = []
    excel_row_number = 1

    for r in rows_iter:
        excel_row_number += 1
        if r is None:
            continue
        values = list(r)

        # fila vacía => skip
        if not any(v is not None and str(v).strip() != "" for v in values):
            continue

        payload: Dict[str, Any] = {}

        for idx, cell in enumerate(values):
            if idx >= len(norm_headers):
                continue
            raw_key = norm_headers[idx]
            if not raw_key:
                continue

            key = header_map.get(raw_key) or raw_key  # si ya viene "km_acum" etc

            # normalizaciones suaves por tipo
            if key in ("codigo", "operador"):
                val = _norm(cell)
            elif key in ("km_acum", "lt_dies_ac"):
                val = _to_float(cell)
            elif key in ("activo", "mexico", "exp_ver", "importado"):
                b = _to_bool(cell)
                val = b if b is not None else cell
            else:
                # deja tal cual (sanitize_car_payload decidirá)
                val = cell

            # si viene vacío, no lo metas
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue

            payload[key] = val

        out.append({"_row_number": excel_row_number, "payload": payload})

    return out


def import_cars_from_excel(
    *,
    client_id: int,
    file_storage: FileStorage,
    dry_run: bool = False,
    upsert_by_code: bool = False,
) -> Dict[str, Any]:
    """
    Lee un Excel y crea/actualiza unidades (cars).

    - codigo: obligatorio (se normaliza upper)
    - si upsert_by_code=1: actualiza por (client_id, codigo)
    - rendimiento_promedio se recalcula (km_acum / lt_dies_ac) usando tu helper
    """
    rows = _read_xlsx_rows(file_storage)
    if not rows:
        raise ApiError("El Excel no tiene filas de datos.", status_code=400)

    created = 0
    updated = 0
    skipped = 0
    errors: List[RowError] = []
    preview: List[Dict[str, Any]] = []

    try:
        for item in rows:
            row_number = item["_row_number"]
            raw_payload = dict(item["payload"])

            codigo = (raw_payload.get("codigo") or "").strip().upper()
            if not codigo:
                errors.append(
                    RowError(
                        row_number=row_number,
                        message="El campo 'codigo' es obligatorio para unidades.",
                        data=raw_payload,
                    )
                )
                continue
            raw_payload["codigo"] = codigo

            # default activo si no viene
            if "activo" not in raw_payload:
                raw_payload["activo"] = True

            # normaliza/valida con tu sanitizer actual
            try:
                data = sanitize_car_payload(raw_payload, partial=False)
            except ApiError as e:
                errors.append(RowError(row_number=row_number, message=str(e), data=raw_payload))
                continue

            data["client_id"] = client_id

            # preview dry-run
            if dry_run:
                preview.append(
                    {
                        "row": row_number,
                        "codigo": data.get("codigo"),
                        "operador": data.get("operador", ""),
                    }
                )
                continue

            if upsert_by_code:
                existing = (
                    Car.query.filter_by(client_id=client_id, codigo=data["codigo"])
                    .limit(1)
                    .first()
                )
                if existing:
                    for k, v in data.items():
                        if k in ("id", "client_id"):
                            continue
                        setattr(existing, k, v)

                    # recalcula rendimiento
                    existing.rendimiento_promedio = calc_rendimiento_promedio(
                        getattr(existing, "km_acum", None),
                        getattr(existing, "lt_dies_ac", None),
                    )
                    updated += 1
                    continue

            car = Car(**data)
            car.rendimiento_promedio = calc_rendimiento_promedio(
                getattr(car, "km_acum", None),
                getattr(car, "lt_dies_ac", None),
            )
            db.session.add(car)
            created += 1

        if dry_run:
            db.session.rollback()
        else:
            db.session.commit()

    except IntegrityError:
        db.session.rollback()
        raise ApiError(
            "Error de integridad al importar unidades (posible código duplicado u otro constraint). "
            "Revisa el archivo y vuelve a intentar.",
            status_code=409,
        )
    except ApiError:
        db.session.rollback()
        raise
    except Exception as e:
        db.session.rollback()
        raise ApiError(f"Error inesperado importando Excel: {e}", status_code=500)

    return {
        "dry_run": dry_run,
        "upsert_by_code": upsert_by_code,
        "total_rows": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors_count": len(errors),
        "errors": [{"row": er.row_number, "message": er.message, "data": er.data} for er in errors],
        "preview": preview if dry_run else [],
    }

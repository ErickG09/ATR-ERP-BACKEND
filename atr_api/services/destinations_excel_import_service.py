from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from werkzeug.datastructures import FileStorage
from sqlalchemy.exc import IntegrityError

from atr_api.extensions import db
from atr_api.errors import ApiError
from atr_api.models.destination import Destination

from openpyxl import load_workbook


@dataclass
class RowError:
    row_number: int
    message: str
    data: Dict[str, Any]


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _norm_header(h: Any) -> str:
    import re
    from unicodedata import normalize

    s = _norm(h).lower()
    s = normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def _to_bool(v: Any) -> Optional[bool]:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("1", "true", "t", "si", "sí", "s", "y", "yes"):
        return True
    if s in ("0", "false", "f", "no", "n"):
        return False
    return None


def _to_num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except Exception:
        return None


def _read_xlsx_rows(file_storage: FileStorage) -> List[Dict[str, Any]]:
    try:
        wb = load_workbook(file_storage.stream, data_only=True)
    except Exception as e:
        raise ApiError(f"No se pudo leer el Excel: {e}", status_code=400)

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    headers = list(headers or [])
    if not any(_norm(h) for h in headers):
        raise ApiError("El Excel no tiene encabezados (fila 1).", status_code=400)

    header_map = {
        "codigo": "codigo",
        "clave": "codigo",
        "destinatario": "nombre",
        "nombre": "nombre",
        "plaza": "plaza",
        "ciudad": "ciudad",
        "estado": "estado",
        "aplicaiva": "aplica_iva",
        "ivapct": "iva_pct",
        "aplicaretencion": "aplica_retencion",
        "retencionpct": "retencion_pct",
        "activo": "activo",
    }

    norm_headers = [_norm_header(h) for h in headers]

    out: List[Dict[str, Any]] = []
    excel_row_number = 1

    for r in rows_iter:
        excel_row_number += 1
        if r is None:
            continue
        values = list(r)
        if not any(v is not None and str(v).strip() != "" for v in values):
            continue

        payload: Dict[str, Any] = {}

        for idx, cell in enumerate(values):
            if idx >= len(norm_headers):
                continue
            raw_key = norm_headers[idx]
            if not raw_key:
                continue

            key = header_map.get(raw_key) or raw_key

            if key in ("codigo", "nombre", "plaza", "ciudad", "estado"):
                val = _norm(cell)
            elif key in ("aplica_iva", "aplica_retencion", "activo"):
                b = _to_bool(cell)
                val = b if b is not None else cell
            elif key in ("iva_pct", "retencion_pct"):
                val = _to_num(cell)
            else:
                val = cell

            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue

            payload[key] = val

        out.append({"_row_number": excel_row_number, "payload": payload})

    return out


def import_destinations_from_excel(
    *,
    client_id: int,
    file_storage: FileStorage,
    dry_run: bool = False,
    upsert_by_code: bool = False,
) -> Dict[str, Any]:
    """
    Lee un Excel y crea/actualiza destinatarios (destinations).

    Reglas (alineadas a tu routes/destinations.py):
    - codigo obligatorio (upper)
    - nombre obligatorio
    - defaults: aplica_iva=True, iva_pct=16, aplica_retencion=False, retencion_pct=0, activo=True
    - upsert_by_code=1 => actualiza por (client_id, codigo)
    """
    rows = _read_xlsx_rows(file_storage)
    if not rows:
        raise ApiError("El Excel no tiene filas de datos.", status_code=400)

    created = 0
    updated = 0
    skipped = 0
    errors: List[RowError] = []
    preview: List[Dict[str, Any]] = []

    try:
        for item in rows:
            row_number = item["_row_number"]
            raw = dict(item["payload"])

            codigo = (raw.get("codigo") or "").strip().upper()
            nombre = (raw.get("nombre") or "").strip()

            if not codigo:
                errors.append(RowError(row_number=row_number, message="El campo 'codigo' es obligatorio.", data=raw))
                continue
            if not nombre:
                errors.append(RowError(row_number=row_number, message="El campo 'nombre' es obligatorio.", data=raw))
                continue

            data = {
                "client_id": client_id,
                "codigo": codigo,
                "nombre": nombre,
                "plaza": (raw.get("plaza") or "").strip() or None,
                "ciudad": (raw.get("ciudad") or "").strip() or None,
                "estado": (raw.get("estado") or "").strip() or None,
                "aplica_iva": bool(raw.get("aplica_iva")) if raw.get("aplica_iva") is not None else True,
                "iva_pct": float(raw.get("iva_pct")) if raw.get("iva_pct") is not None else 16.0,
                "aplica_retencion": bool(raw.get("aplica_retencion")) if raw.get("aplica_retencion") is not None else False,
                "retencion_pct": float(raw.get("retencion_pct")) if raw.get("retencion_pct") is not None else 0.0,
                "activo": bool(raw.get("activo")) if raw.get("activo") is not None else True,
            }

            if dry_run:
                preview.append(
                    {"row": row_number, "codigo": data["codigo"], "nombre": data["nombre"]}
                )
                continue

            if upsert_by_code:
                existing = (
                    Destination.query.filter_by(client_id=client_id, codigo=data["codigo"])
                    .limit(1)
                    .first()
                )
                if existing:
                    for k, v in data.items():
                        if k in ("id", "client_id", "codigo"):
                            continue
                        setattr(existing, k, v)
                    updated += 1
                    continue

            d = Destination(**data)
            db.session.add(d)
            created += 1

        if dry_run:
            db.session.rollback()
        else:
            db.session.commit()

    except IntegrityError:
        db.session.rollback()
        raise ApiError(
            "Error de integridad al importar destinatarios (posible código duplicado u otro constraint). "
            "Revisa el archivo y vuelve a intentar.",
            status_code=409,
        )
    except ApiError:
        db.session.rollback()
        raise
    except Exception as e:
        db.session.rollback()
        raise ApiError(f"Error inesperado importando Excel: {e}", status_code=500)

    return {
        "dry_run": dry_run,
        "upsert_by_code": upsert_by_code,
        "total_rows": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors_count": len(errors),
        "errors": [{"row": er.row_number, "message": er.message, "data": er.data} for er in errors],
        "preview": preview if dry_run else [],
    }

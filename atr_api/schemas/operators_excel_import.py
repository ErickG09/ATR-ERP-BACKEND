from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from typing import Any, Dict, List, Callable

from werkzeug.datastructures import FileStorage

from atr_api.errors import ApiError
from atr_api.models import Operator
from atr_api.extensions import db

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


def _strip_accents(value: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", value) if not unicodedata.combining(c)
    )


def _norm_header(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = _strip_accents(s)
    s = s.replace("_", " ")
    s = re.sub(r"[^\w\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_full_name(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _parse_bool_si_no(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value

    s = _strip_accents(str(value).strip().lower())
    if s in ("1", "true", "t", "yes", "y", "si", "s", "activo"):
        return True
    if s in ("0", "false", "f", "no", "n", "", "inactivo"):
        return False

    raise ApiError("El campo 'Seguro' debe ser SI/NO, T/F, true/false o 1/0.", 400)


def _parse_date_flexible(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)):
        try:
            d = from_excel(value)
            if isinstance(d, datetime):
                return d.date().isoformat()
            if isinstance(d, date):
                return d.isoformat()
        except Exception:
            pass

    s = str(value).strip()
    if not s:
        return None

    try:
        return date.fromisoformat(s).isoformat()
    except Exception:
        pass

    m = re.match(r"^(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})$", s)
    if m:
        dd = int(m.group(1))
        mm = int(m.group(2))
        yyyy = int(m.group(3))
        try:
            return date(yyyy, mm, dd).isoformat()
        except Exception:
            raise ApiError("Fecha inválida en Excel. Usa YYYY-MM-DD o una fecha válida.", 400)

    raise ApiError("Fecha inválida. Usa YYYY-MM-DD o una fecha válida de Excel.", 400)


def _parse_numeric_or_blank(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str) and not value.strip():
        return ""
    return value


def _resolve_header_field(header: str) -> str | None:
    """
    Convierte encabezados variados del Excel a un campo backend.
    Soporta variantes como:
    - Man. Nac / Man Nac / Man.Nac
    - Ma. Ramos-Altamira / Man Ramos Altamira
    - Man.Slp-Lzc / Man Slp Lzc
    """
    h = _norm_header(header)
    if not h:
        return None

    direct_map: Dict[str, str] = {
        # identificadores
        "codigo": "codigo",
        "codigo operador": "codigo",
        "nombre": "nombre",
        "fecha ingreso": "fecha_ingreso",
        "estatus": "activo",
        "status": "activo",

        # sueldos / viáticos
        "sueldo op1": "sueldo_op_1",
        "sueldo operador 1": "sueldo_op_1",
        "viaticos op1": "viaticos_op_1",
        "viaticos operador 1": "viaticos_op_1",
        "sueldo op2": "sueldo_op_2",
        "sueldo operador 2": "sueldo_op_2",
        "viaticos op2": "viaticos_op_2",
        "viaticos operador 2": "viaticos_op_2",
        "viaje esp": "viaje_especial",
        "viaje especial": "viaje_especial",

        # tarifas
        "d f": "mexico",
        "df": "mexico",
        "mexico": "mexico",
        "exp ver": "exp_ver",
        "exp lzc": "exp_lc",
        "exp lc": "exp_lc",
        "exp tux": "exp_tux",
        "importado": "importado",
        "local": "local",
        "patios": "patios",
        "slp altamira": "slp_altamira",
        "ramos altamira": "ramos_altamira",
        "slp lzc": "slp_lc",
        "slp lc": "slp_lc",
        "salamanca lzc": "sal_lzc",
        "salamanca lc": "sal_lzc",
        "salamanca ver": "sal_ver",
        "salamanca altamira": "sal_altamira",
        "resguardo": "resguardo",
        "ayuda esc": "ayuda_escolar",
        "ayuda escolar": "ayuda_escolar",

        # datos personales
        "domicilio": "domicilio",
        "imss": "no_imss",
        "no imss": "no_imss",
        "nss": "no_imss",
        "licencia": "no_licencia",
        "no licencia": "no_licencia",
        "vence licencia": "fecha_venc_licencia",
        "vencimiento licencia": "fecha_venc_licencia",
        "telefono": "telefono",
        "rfc": "rfc",
        "email": "correo_electronico",
        "correo": "correo_electronico",
        "correo electronico": "correo_electronico",
        "gafete aduana": "gafete_aduana",
        "gafete": "gafete_aduana",
        "apto medico": "apto_medico_licencia",
        "seguro": "tiene_seguro",
        "tipo carro": "tipo_carro",
        "tipo de carro": "tipo_carro",
        "tipo unidad": "tipo_carro",
    }

    if h in direct_map:
        return direct_map[h]

    tokens = h.split()

    # Maniobras: tolera "man", "ma", "maniobra", "maniobras"
    if tokens and tokens[0] in {"man", "ma", "maniobra", "maniobras"}:
        route = " ".join(tokens[1:]).strip()

        maniobra_map = {
            "nac": "man_nac",
            "nacional": "man_nac",
            "esp": "man_esp",
            "especial": "man_esp",
            "d f": "man_df",
            "df": "man_df",
            "ver": "man_ver",
            "veracruz": "man_ver",
            "slp altamira": "man_slp_altamira",
            "san luis altamira": "man_slp_altamira",
            "ramos altamira": "man_ramos_altamira",
            "slp lzc": "man_slp_lzc",
            "slp lc": "man_slp_lzc",
            "san luis lzc": "man_slp_lzc",
            "salamanca lzc": "man_salamanca_lzc",
            "salamanca lc": "man_salamanca_lzc",
            "salamanca ver": "man_salamanca_ver",
            "salamanca altamira": "man_salamanca_altamira",
        }

        if route in maniobra_map:
            return maniobra_map[route]

    return None


def parse_excel_operators(file_storage: FileStorage) -> List[Dict[str, Any]]:
    try:
        file_storage.stream.seek(0)
        wb = load_workbook(file_storage.stream, data_only=True)
    except Exception as e:
        raise ApiError(f"Excel inválido o corrupto: {e}", 400)

    ws = wb.active

    raw_headers: List[Any] = [c.value for c in ws[1]]
    headers = [_norm_header(v) for v in raw_headers]

    if not any(headers):
        raise ApiError("No se detectaron encabezados en la fila 1.", 400)

    col_to_field: Dict[int, str] = {}
    for idx, raw_header in enumerate(raw_headers):
        field = _resolve_header_field(raw_header)
        if field:
            col_to_field[idx] = field

    if "nombre" not in col_to_field.values():
        raise ApiError("El Excel debe incluir la columna 'Nombre'.", 400)

    rows_out: List[Dict[str, Any]] = []

    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        payload: Dict[str, Any] = {}

        for col_idx, field in col_to_field.items():
            value = row[col_idx].value if col_idx < len(row) else None

            if field in ("fecha_ingreso", "fecha_venc_licencia", "apto_medico_licencia"):
                payload[field] = _parse_date_flexible(value)
            elif field in ("tiene_seguro", "activo"):
                payload[field] = _parse_bool_si_no(value)
            elif field in {
                "sueldo_op_1",
                "viaticos_op_1",
                "sueldo_op_2",
                "viaticos_op_2",
                "viaje_especial",
                "mexico",
                "exp_ver",
                "exp_lc",
                "exp_tux",
                "importado",
                "local",
                "patios",
                "slp_altamira",
                "ramos_altamira",
                "slp_lc",
                "sal_lzc",
                "sal_ver",
                "sal_altamira",
                "resguardo",
                "ayuda_escolar",
                "man_nac",
                "man_esp",
                "man_df",
                "man_ver",
                "man_slp_altamira",
                "man_ramos_altamira",
                "man_slp_lzc",
                "man_salamanca_lzc",
                "man_salamanca_ver",
                "man_salamanca_altamira",
            }:
                payload[field] = _parse_numeric_or_blank(value)
            else:
                payload[field] = "" if value is None else str(value).strip()

        if not normalize_full_name(payload.get("nombre", "")):
            continue

        rows_out.append({"_row_number": row_idx, "payload": payload})

    return rows_out


_CODE_RE = re.compile(r"^([A-Z])(\d+)$")


def _clean_initial(letter: str) -> str:
    letter = (letter or "").strip().upper()
    return letter[0] if letter else "X"


def _extract_prefix_from_name(full_name: str) -> str:
    parts = normalize_full_name(full_name).split(" ")
    first = parts[0] if parts else ""
    return _clean_initial(first[:1])


def _normalize_codigo(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().upper()
    s = re.sub(r"\s+", "", s)
    return s


def make_operator_code_generator(*, client_id: int) -> Callable[[str], str]:
    cache_used: Dict[str, set[int]] = {}
    cache_next_candidate: Dict[str, int] = {}

    def _load_used(prefix: str) -> set[int]:
        if prefix in cache_used:
            return cache_used[prefix]

        like = f"{prefix}%"
        existing = (
            db.session.query(Operator.codigo)
            .filter(Operator.client_id == client_id)
            .filter(Operator.codigo.ilike(like))
            .all()
        )

        used: set[int] = set()
        for (code,) in existing:
            if not code:
                continue
            m = _CODE_RE.match(str(code).strip().upper())
            if not m:
                continue
            try:
                used.add(int(m.group(2)))
            except Exception:
                continue

        cache_used[prefix] = used
        cache_next_candidate[prefix] = 1
        return used

    def _next_available(prefix: str) -> int:
        used = _load_used(prefix)
        n = cache_next_candidate.get(prefix, 1)

        while n in used:
            n += 1

        used.add(n)
        cache_next_candidate[prefix] = n + 1
        return n

    def generator(*, full_name: str) -> str:
        prefix = _extract_prefix_from_name(full_name)
        n = _next_available(prefix)
        return f"{prefix}{n:03d}"

    return generator


def normalize_codigo_from_excel(*, codigo: Any, nombre: Any) -> str:
    c = _normalize_codigo(codigo)
    if not c:
        return ""

    m = _CODE_RE.match(c)
    if not m:
        raise ApiError(
            f"Código inválido '{c}'. Debe tener formato como A006 (letra + número).",
            400,
        )

    prefix = m.group(1)
    num = int(m.group(2))
    return f"{prefix}{num:03d}"
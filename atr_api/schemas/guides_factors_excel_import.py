# atr_api/schemas/guides_factors_excel_import.py
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from werkzeug.datastructures import FileStorage

from openpyxl import load_workbook

from atr_api.errors import ApiError


# -----------------------------------------------------------------------------
# Helpers de normalización
# -----------------------------------------------------------------------------
def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )
    s = re.sub(r"\s+", " ", s)
    s = s.replace("\n", " ").replace("\r", " ")
    return s.strip()


def _to_str(v: Any) -> str:
    return ("" if v is None else str(v)).strip()


def _to_upper(v: Any) -> str:
    return _to_str(v).upper()


def _parse_int_strict(v: Any) -> Optional[int]:
    """
    Entero estricto (sin decimales).
    Acepta: 280, "280"
    Rechaza: 280.0, "280.0", "280,0", "280.5"
    """
    if v is None or v == "":
        return None

    if isinstance(v, bool):
        return None

    if isinstance(v, int):
        return v

    if isinstance(v, float):
        # aunque sea 280.0, lo rechazamos para ser estrictos
        return None

    s = _to_str(v)
    if not s:
        return None

    if "." in s or "," in s:
        return None

    try:
        return int(s)
    except Exception:
        return None


_CURRENCY_RE = re.compile(r"[^\d\-\.,]")


def _parse_money(v: Any) -> Optional[Decimal]:
    """
    Parsea importes tipo "$20,641.60" / "20641.60" / "20,641.60".
    Regla:
    - Si hay '.' y ',' => asumimos ',' miles y '.' decimal -> quitamos ','.
    - Si solo hay ',' => asumimos ',' decimal -> reemplazamos por '.'.
    """
    if v is None or v == "":
        return None

    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except Exception:
            return None

    s = _to_str(v)
    if not s:
        return None

    # quita $ y letras
    s = _CURRENCY_RE.sub("", s)

    if not s:
        return None

    if "," in s and "." in s:
        s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")

    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


# -----------------------------------------------------------------------------
# Encabezados soportados (aliases)
# -----------------------------------------------------------------------------
# Mapeo: "encabezado en Excel normalizado" -> "campo interno"
HEADER_ALIASES: Dict[str, str] = {
    "carro": "carro",
    "unidad": "carro",
    "tipo de carro": "carro",
    "tipo carro": "carro",
    "tipo unidad": "carro",
    "factor": "importe",          # en tu excel se llama "FACTOR" pero es la tarifa/importe
    "importe": "importe",
    "tarifa": "importe",
    "monto": "importe",
    "kms": "kms",
    "km": "kms",
    "kilometros": "kms",
    "kilómetros": "kms",
    "td": "td",
    "tipo dest": "td",
    "tipo destino": "td",
}


REQUIRED_FIELDS = ("carro", "td", "kms", "importe")


@dataclass
class FactorRowError:
    row_number: int
    message: str
    data: Dict[str, Any]


def _pick_sheet_name(wb) -> str:
    """
    FACTORES.xls que compartiste trae sheet "CAJA".
    Si no existe, tomamos la primera hoja.
    """
    wanted = {"CAJA", "FACTORES", "FACTORS"}
    for name in wb.sheetnames:
        if name.strip().upper() in wanted:
            return name
    return wb.sheetnames[0]


def _read_header_map(header_row: List[Any]) -> Dict[int, str]:
    """
    Devuelve un mapping: index_col (0-based) -> campo interno
    Solo para columnas reconocidas.
    """
    col_map: Dict[int, str] = {}
    for idx, cell_val in enumerate(header_row):
        key = _norm_header(_to_str(cell_val))
        if not key:
            continue
        internal = HEADER_ALIASES.get(key)
        if internal:
            col_map[idx] = internal
    return col_map


def parse_excel_guide_factors(file: FileStorage) -> Dict[str, Any]:
    """
    Lee un Excel de FACTORES y regresa:
      {
        "ok": bool,
        "sheet": str,
        "rows": [ {carro, td, kms, importe, ...} ],
        "errors": [ {row_number, message, data} ],
        "counts": {"total": int, "valid": int, "invalid": int}
      }

    Nota importante:
    - openpyxl NO soporta .xls (formato binario antiguo). Si tu archivo es .xls real,
      aquí fallará. En ese caso, conviértelo a .xlsx.
    """
    if not file:
        raise ApiError("Archivo requerido.", status_code=400)

    filename = (getattr(file, "filename", "") or "").lower()
    if not filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")):
        raise ApiError("Formato inválido. Sube un Excel (.xlsx).", status_code=400)

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        # Muy común: .xls real (binario). openpyxl no lo abre.
        raise ApiError(
            "No pude leer el Excel. Si tu archivo es .xls (formato antiguo), "
            "conviértelo a .xlsx y vuelve a intentarlo.",
            status_code=400,
        ) from e

    sheet_name = _pick_sheet_name(wb)
    ws = wb[sheet_name]

    # Busca la primera fila que parezca header (contenga al menos 2 columnas reconocidas)
    header_row_idx: Optional[int] = None
    header_map: Dict[int, str] = {}
    max_scan = min(ws.max_row or 0, 50)

    for r in range(1, max_scan + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, (ws.max_column or 1) + 1)]
        candidate_map = _read_header_map(row_vals)
        if len(candidate_map) >= 2:
            header_row_idx = r
            header_map = candidate_map
            break

    if header_row_idx is None:
        raise ApiError(
            "No se encontraron encabezados válidos. Se esperan columnas como: "
            "CARRO, TD, KMS, FACTOR.",
            status_code=400,
        )

    # Valida que estén los campos mínimos
    found_fields = set(header_map.values())
    missing = [f for f in REQUIRED_FIELDS if f not in found_fields]
    if missing:
        raise ApiError(
            f"Faltan columnas requeridas en FACTORES: {', '.join(missing)}.",
            status_code=400,
        )

    rows: List[Dict[str, Any]] = []
    errors: List[FactorRowError] = []
    total = 0

    # Recorre filas de datos
    for r in range(header_row_idx + 1, (ws.max_row or header_row_idx) + 1):
        raw: Dict[str, Any] = {}
        # extrae solo columnas conocidas
        for col_idx, field in header_map.items():
            raw[field] = ws.cell(row=r, column=col_idx + 1).value

        # Detecta fila vacía (de las columnas que importan)
        if all(_to_str(raw.get(f)) == "" for f in REQUIRED_FIELDS):
            continue

        total += 1
        out: Dict[str, Any] = {}

        carro = _to_upper(raw.get("carro"))
        td = _to_upper(raw.get("td"))

        kms = _parse_int_strict(raw.get("kms"))
        importe = _parse_money(raw.get("importe"))

        if not carro:
            errors.append(FactorRowError(r, "CARRO vacío.", raw))
            continue
        if not td:
            errors.append(FactorRowError(r, "TD vacío.", raw))
            continue
        if kms is None:
            errors.append(FactorRowError(r, "KMS inválido (debe ser entero, sin decimales).", raw))
            continue
        if kms < 0:
            errors.append(FactorRowError(r, "KMS debe ser >= 0.", raw))
            continue
        if importe is None:
            errors.append(FactorRowError(r, "FACTOR/IMPORTE inválido.", raw))
            continue
        if importe < 0:
            errors.append(FactorRowError(r, "FACTOR/IMPORTE debe ser >= 0.", raw))
            continue

        out["carro"] = carro
        out["td"] = td
        out["kms"] = int(kms)
        # guardamos como string para evitar problemas JSON con Decimal; el service decide
        out["importe"] = str(importe.quantize(Decimal("0.01")))

        rows.append(out)

    payload = {
        "ok": len(errors) == 0,
        "sheet": sheet_name,
        "rows": rows,
        "errors": [
            {"row_number": e.row_number, "message": e.message, "data": e.data}
            for e in errors
        ],
        "counts": {"total": total, "valid": len(rows), "invalid": len(errors)},
    }
    return payload
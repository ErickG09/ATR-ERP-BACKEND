# atr_api/schemas/liquidaciones_excel_import.py

from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from werkzeug.datastructures import FileStorage

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from atr_api.errors import ApiError


# -----------------------------------------------------------------------------
# Encabezados soportados (aliases)
# -----------------------------------------------------------------------------
HEADER_ALIASES_RAW: Dict[str, str] = {
    # Talón / viaje
    "talon": "talon_interno",
    "talón": "talon_interno",
    "talon interno": "talon_interno",
    "talón interno": "talon_interno",
    "talon/viaje": "talon_interno",
    "talón/viaje": "talon_interno",
    "talon viaje": "talon_interno",
    "talón viaje": "talon_interno",
    "talon-viaje": "talon_interno",
    "talón-viaje": "talon_interno",
    "viaje": "talon_interno",
    "no viaje": "talon_interno",
    "n° viaje": "talon_interno",
    "nº viaje": "talon_interno",
    "num viaje": "talon_interno",
    "numero viaje": "talon_interno",
    "n viaje": "talon_interno",

    # Fecha
    "fecha": "fecha",
    "fecha viaje": "fecha",

    # Operadores (tu Excel)
    "operador": "operador_1",
    "ayudante": "operador_2",
    "chofer": "operador_1",

    # Formatos alternativos
    "1er operador": "operador_1",
    "1 operador": "operador_1",
    "operador 1": "operador_1",
    "primer operador": "operador_1",

    "2º operador": "operador_2",
    "2o operador": "operador_2",
    "2do operador": "operador_2",
    "2 operador": "operador_2",
    "operador 2": "operador_2",
    "segundo operador": "operador_2",

    # Factura / Carta Porte
    "factura": "factura_cp",
    "c.p.": "factura_cp",
    "c.p": "factura_cp",
    "cp": "factura_cp",
    "carta porte": "factura_cp",
    "factura/c.p.": "factura_cp",
    "factura/c.p": "factura_cp",
    "factura/carta porte": "factura_cp",
    "factura/cp": "factura_cp",
    "factura cp": "factura_cp",
    "factura c p": "factura_cp",

    # Carro / unidad (camión) / placas
    "carro": "carro",
    "unidad": "carro",
    "camion": "carro",
    "camión": "carro",
    "tracto": "carro",
    "placas": "carro",

    # Dealer / destino
    "dealer": "dealer",
    "agencia": "dealer",
    "destino": "dealer",
    "destinatario": "dealer",

    # Unidades
    "unidades": "unidades",
    "uds": "unidades",
    "cantidad": "unidades",
    "cantidad unidades": "unidades",

    # Kms
    "kms": "kms",
    "km": "kms",
    "kilometros": "kms",
    "kilómetros": "kms",

    # Flete / subtotal
    "flete": "flete",
    "subtotal": "flete",

    # IVA / retención / total
    "iva": "iva",
    "retencion": "retencion",
    "retención": "retencion",
    "total": "total",

    # Anticipos / recibos
    "anticipo de gastos 1º": "anticipo_1",
    "anticipo de gastos 1": "anticipo_1",
    "anticipo gastos 1": "anticipo_1",
    "anticipo 1": "anticipo_1",
    "anticipo operador 1": "anticipo_1",
    "anticipo de gastos 1o": "anticipo_1",

    "folio recibo 1º": "recibo_1",
    "folio recibo 1": "recibo_1",
    "recibo 1": "recibo_1",
    "folio 1": "recibo_1",

    "anticipo de gastos 2º": "anticipo_2",
    "anticipo de gastos 2": "anticipo_2",
    "anticipo gastos 2": "anticipo_2",
    "anticipo 2": "anticipo_2",
    "anticipo operador 2": "anticipo_2",
    "anticipo de gastos 2o": "anticipo_2",

    "folio recibo 2º": "recibo_2",
    "folio recibo 2": "recibo_2",
    "recibo 2": "recibo_2",
    "folio 2": "recibo_2",
}


# -----------------------------------------------------------------------------
# Normalización y parseos básicos
# -----------------------------------------------------------------------------
def _strip_accents(s: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch)
    )


def _norm_header(s: Any) -> str:
    """
    Normaliza encabezados:
      - lower
      - sin acentos
      - colapsa espacios
      - reemplaza NBSP, tabs
      - convierte . / - a espacios
    """
    if s is None:
        return ""
    raw = str(s)
    raw = raw.replace("\xa0", " ")  # NBSP
    raw = raw.replace("\t", " ")
    raw = raw.strip().lower()
    raw = raw.replace("\n", " ").replace("\r", " ")
    raw = _strip_accents(raw)

    raw = raw.replace(".", " ").replace("/", " ").replace("-", " ")
    raw = re.sub(r"\s+", " ", raw).strip()
    return raw


def _header_compact_key(h_norm: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", h_norm or "")


def _clean_string(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).replace("\xa0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _parse_date_flexible(value: Any) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)):
        try:
            d = from_excel(value)
            if isinstance(d, datetime):
                return d.date().isoformat()
            if isinstance(d, date):
                return d.isoformat()
        except Exception:
            pass

    s = _clean_string(value)
    if not s:
        return None

    try:
        return date.fromisoformat(s).isoformat()
    except Exception:
        pass

    m = re.match(r"^(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})$", s)
    if m:
        dd = int(m.group(1))
        mm = int(m.group(2))
        yyyy = int(m.group(3))
        try:
            return date(yyyy, mm, dd).isoformat()
        except Exception:
            raise ApiError("Fecha inválida en Excel. Usa YYYY-MM-DD o una fecha válida.", 400)

    raise ApiError("Fecha inválida. Usa YYYY-MM-DD o una fecha válida de Excel.", 400)


def _parse_int(value: Any, *, allow_blank: bool = True) -> Optional[int]:
    if value is None or isinstance(value, bool):
        return None if allow_blank else 0
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float):
        return int(value) if value.is_integer() else int(round(value))

    s = _clean_string(value)
    if not s:
        return None if allow_blank else 0

    s2 = re.sub(r"[^\d\-]+", "", s)
    if not s2 or s2 in ("-",):
        return None if allow_blank else 0

    try:
        return int(s2)
    except Exception:
        raise ApiError(f"Entero inválido: '{s}'", 400)


def _parse_money(value: Any, *, allow_blank: bool = True) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None if allow_blank else 0.0
    if isinstance(value, (int, float)):
        return float(value)

    s = _clean_string(value)
    if not s:
        return None if allow_blank else 0.0

    s = s.replace("$", "").replace(" ", "")

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        if "," in s and "." not in s:
            s = s.replace(",", ".")

    s = re.sub(r"[^0-9\.\-]+", "", s)
    if not s or s in ("-", ".", "-."):
        return None if allow_blank else 0.0

    try:
        return float(s)
    except Exception:
        raise ApiError(f"Numérico inválido: '{value}'", 400)


def _normalize_talon(value: Any) -> str:
    if value is None:
        return ""
    return "".join(_clean_string(value).split()).upper()


def _build_header_aliases_normalized() -> Dict[str, str]:
    out: Dict[str, str] = {}
    for k, v in HEADER_ALIASES_RAW.items():
        nk = _norm_header(k)
        if nk:
            out[nk] = v
    return out


HEADER_ALIASES: Dict[str, str] = _build_header_aliases_normalized()
HEADER_ALIASES_COMPACT: Dict[str, str] = {
    _header_compact_key(_norm_header(k)): v
    for k, v in HEADER_ALIASES_RAW.items()
    if _header_compact_key(_norm_header(k))
}


def _resolve_field_from_header(h_norm: str) -> Optional[str]:
    if not h_norm:
        return None

    if h_norm in HEADER_ALIASES:
        return HEADER_ALIASES[h_norm]

    ck = _header_compact_key(h_norm)
    if ck in HEADER_ALIASES_COMPACT:
        return HEADER_ALIASES_COMPACT[ck]

    # Último recurso: heurísticas 1/2 para anticipo/recibo
    if "anticipo" in h_norm and re.search(r"\b1\b", h_norm):
        return "anticipo_1"
    if "anticipo" in h_norm and re.search(r"\b2\b", h_norm):
        return "anticipo_2"
    if ("folio" in h_norm or "recibo" in h_norm) and re.search(r"\b1\b", h_norm):
        return "recibo_1"
    if ("folio" in h_norm or "recibo" in h_norm) and re.search(r"\b2\b", h_norm):
        return "recibo_2"

    return None


# -----------------------------------------------------------------------------
# Detección automática de fila de encabezados
# -----------------------------------------------------------------------------
HEADER_SCAN_MAX_ROWS = 15


def _build_col_to_field_from_row(values: Tuple[Any, ...]) -> Dict[int, str]:
    col_to_field: Dict[int, str] = {}
    seen_fields: set[str] = set()

    for idx, cell_value in enumerate(values):
        h_norm = _norm_header(cell_value)
        field = _resolve_field_from_header(h_norm)
        if not field:
            continue

        # Si hay duplicados en encabezado (ej: "TOTAL" repetido), no sobrescribas.
        if field in seen_fields:
            continue

        col_to_field[idx] = field
        seen_fields.add(field)

    return col_to_field


def _detect_header_row(ws) -> Tuple[int, Dict[int, str]]:
    best_row = 1
    best_map: Dict[int, str] = {}
    best_score = -1
    best_has_talon = False

    max_row = min(ws.max_row or 1, HEADER_SCAN_MAX_ROWS)

    for r in range(1, max_row + 1):
        row_values = tuple(c.value for c in ws[r])
        col_to_field = _build_col_to_field_from_row(row_values)

        recognized_fields = set(col_to_field.values())
        score = len(recognized_fields)
        has_talon = "talon_interno" in recognized_fields

        if (
            score > best_score
            or (score == best_score and has_talon and not best_has_talon)
            or (score == best_score and has_talon == best_has_talon and r < best_row)
        ):
            best_row = r
            best_map = col_to_field
            best_score = score
            best_has_talon = has_talon

    if best_score <= 0:
        raise ApiError(
            "No se pudieron detectar encabezados válidos en las primeras filas del Excel. "
            "Verifica que existan columnas como 'TALON', 'Fecha', 'Operador', etc.",
            400,
        )

    return best_row, best_map


def _pick_worksheet(wb):
    """
    Si tu archivo tiene varias hojas, esto reduce el riesgo de leer la equivocada.
    Regla:
      - si existe una hoja cuyo nombre contenga 'liq' o 'liquid' o 'viaje', úsala
      - si no, usa wb.active
    """
    candidates = []
    for name in wb.sheetnames:
        n = _norm_header(name)
        if "liq" in n or "liquid" in n or "viaje" in n:
            candidates.append(name)

    if candidates:
        return wb[candidates[0]]

    return wb.active


# -----------------------------------------------------------------------------
# Parser principal
# -----------------------------------------------------------------------------
def parse_excel_liquidaciones(file_storage: FileStorage) -> List[Dict[str, Any]]:
    """
    Regresa:
      [{ "_row_number": <fila_excel>, "payload": {...}}, ...]
    """
    try:
        file_storage.stream.seek(0)
        wb = load_workbook(file_storage.stream, data_only=True)
    except Exception as e:
        raise ApiError(f"Excel inválido o corrupto: {e}", 400)

    ws = _pick_worksheet(wb)

    header_row_idx, col_to_field = _detect_header_row(ws)

    if "talon_interno" not in col_to_field.values():
        raise ApiError(
            "No se encontró la columna del talón (ej. 'TALON', 'TALON/VIAJE', 'Talón interno'). "
            f"Se detectó encabezado en fila {header_row_idx}, pero sin talón.",
            400,
        )

    rows_out: List[Dict[str, Any]] = []

    for row_idx in range(header_row_idx + 1, (ws.max_row or header_row_idx) + 1):
        row = ws[row_idx]
        payload: Dict[str, Any] = {}

        for col_idx, field in col_to_field.items():
            value = row[col_idx].value if col_idx < len(row) else None

            if field == "talon_interno":
                payload[field] = _normalize_talon(value)
            elif field == "fecha":
                payload[field] = _parse_date_flexible(value)
            elif field == "unidades":
                payload[field] = _parse_int(value, allow_blank=True)
            elif field == "kms":
                payload[field] = _parse_money(value, allow_blank=True)
            elif field in ("flete", "iva", "retencion", "total", "anticipo_1", "anticipo_2"):
                payload[field] = _parse_money(value, allow_blank=True)
            else:
                payload[field] = _clean_string(value)

        talon = _normalize_talon(payload.get("talon_interno", ""))
        if not talon:
            continue
        payload["talon_interno"] = talon

        # Importante: conservar fila original para ordenar después en el service
        rows_out.append({"_row_number": row_idx, "payload": payload})

    if not rows_out:
        raise ApiError(
            "El Excel no tiene filas de datos (o todas venían sin talón) después de la fila de encabezados.",
            400,
        )

    return rows_out
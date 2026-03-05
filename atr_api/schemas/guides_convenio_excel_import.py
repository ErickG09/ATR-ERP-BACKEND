# atr_api/schemas/guides_convenio_excel_import.py
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from werkzeug.datastructures import FileStorage

from openpyxl import load_workbook

from atr_api.errors import ApiError


# -----------------------------------------------------------------------------
# Helpers de normalización
# -----------------------------------------------------------------------------
def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )
    s = re.sub(r"\s+", " ", s)
    s = s.replace("\n", " ").replace("\r", " ")
    return s.strip()


def _to_str(v: Any) -> str:
    return ("" if v is None else str(v)).strip()


def _to_upper(v: Any) -> str:
    return _to_str(v).upper()


def _parse_int_strict(v: Any) -> Optional[int]:
    """
    Entero estricto (sin decimales).
    Acepta: 48, "48"
    Rechaza: 48.0, "48.0", "48,0", "48.5"
    """
    if v is None or v == "":
        return None

    if isinstance(v, bool):
        return None

    if isinstance(v, int):
        return v

    if isinstance(v, float):
        return None

    s = _to_str(v)
    if not s:
        return None

    if "." in s or "," in s:
        return None

    try:
        return int(s)
    except Exception:
        return None


def _parse_codigo(v: Any, pad_left: int = 4) -> Optional[str]:
    """
    Normaliza la CLAVE del destinatario.
    - Si viene numérico (1) => "0001" (por default pad_left=4)
    - Si viene string ("0001") => "0001"
    - Si viene con espacios => trim
    - Si viene con .0 (Excel) => se rechaza (para evitar ambigüedad) y que lo arreglen en el Excel
      (si prefieres aceptarlo, podemos permitir float entero, pero tú pediste estricto)
    """
    if v is None or v == "":
        return None

    if isinstance(v, bool):
        return None

    if isinstance(v, int):
        s = str(v)
        return s.zfill(pad_left) if pad_left and len(s) < pad_left else s

    if isinstance(v, float):
        # estricto: no aceptamos floats aunque sean .0
        return None

    s = _to_str(v)
    if not s:
        return None

    # si trae decimales, lo rechazamos
    if "." in s or "," in s:
        return None

    # si es numérico en string, pad
    if s.isdigit():
        return s.zfill(pad_left) if pad_left and len(s) < pad_left else s

    # si no es dígitos puros, igual lo aceptamos como string "código" (ej. "A12")
    # lo normalizamos a upper por consistencia
    return s.upper()


# -----------------------------------------------------------------------------
# Encabezados soportados (aliases)
# -----------------------------------------------------------------------------
# Mapeo: "encabezado en Excel normalizado" -> "campo interno"
HEADER_ALIASES: Dict[str, str] = {
    "clave": "destination_codigo",
    "codigo": "destination_codigo",
    "cve": "destination_codigo",
    "cvdest": "destination_codigo",
    "cv. dest": "destination_codigo",
    "cv dest": "destination_codigo",
    "cod destinatario": "destination_codigo",
    "kms": "kms",
    "km": "kms",
    "kilometros": "kms",
    "kilómetros": "kms",
    "tipo dest": "td",
    "tipo destino": "td",
    "td": "td",
    "destinat": "destinatario_nombre",
    "destinatario": "destinatario_nombre",
    "nombre": "destinatario_nombre",
    "ciudad": "ciudad",
}


REQUIRED_FIELDS = ("destination_codigo", "kms", "td")


@dataclass
class ConvenioRowError:
    row_number: int
    message: str
    data: Dict[str, Any]


def _pick_sheet_name(wb) -> str:
    """
    CONVENIO.xls que compartiste trae sheet "CONVENIO".
    Si no existe, tomamos la primera hoja.
    """
    wanted = {"CONVENIO", "CONVENIOS"}
    for name in wb.sheetnames:
        if name.strip().upper() in wanted:
            return name
    return wb.sheetnames[0]


def _read_header_map(header_row: List[Any]) -> Dict[int, str]:
    """
    Devuelve un mapping: index_col (0-based) -> campo interno
    Solo para columnas reconocidas.
    """
    col_map: Dict[int, str] = {}
    for idx, cell_val in enumerate(header_row):
        key = _norm_header(_to_str(cell_val))
        if not key:
            continue
        internal = HEADER_ALIASES.get(key)
        if internal:
            col_map[idx] = internal
    return col_map


def parse_excel_guide_convenio(
    file: FileStorage,
    *,
    codigo_pad_left: int = 4,
) -> Dict[str, Any]:
    """
    Lee un Excel de CONVENIO y regresa:
      {
        "ok": bool,
        "sheet": str,
        "rows": [ {destination_codigo, kms, td, destinatario_nombre?, ciudad?} ],
        "errors": [ {row_number, message, data} ],
        "counts": {"total": int, "valid": int, "invalid": int},
        "settings": {"codigo_pad_left": int}
      }

    Nota importante:
    - openpyxl NO soporta .xls (formato binario antiguo). Si tu archivo es .xls real,
      aquí fallará. En ese caso, conviértelo a .xlsx.
    """
    if not file:
        raise ApiError("Archivo requerido.", status_code=400)

    filename = (getattr(file, "filename", "") or "").lower()
    if not filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")):
        raise ApiError("Formato inválido. Sube un Excel (.xlsx).", status_code=400)

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        raise ApiError(
            "No pude leer el Excel. Si tu archivo es .xls (formato antiguo), "
            "conviértelo a .xlsx y vuelve a intentarlo.",
            status_code=400,
        ) from e

    sheet_name = _pick_sheet_name(wb)
    ws = wb[sheet_name]

    # Busca header
    header_row_idx: Optional[int] = None
    header_map: Dict[int, str] = {}
    max_scan = min(ws.max_row or 0, 50)

    for r in range(1, max_scan + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, (ws.max_column or 1) + 1)]
        candidate_map = _read_header_map(row_vals)
        if len(candidate_map) >= 2:
            header_row_idx = r
            header_map = candidate_map
            break

    if header_row_idx is None:
        raise ApiError(
            "No se encontraron encabezados válidos. Se esperan columnas como: "
            "CLAVE, KMS, TIPO DEST.",
            status_code=400,
        )

    found_fields = set(header_map.values())
    missing = [f for f in REQUIRED_FIELDS if f not in found_fields]
    if missing:
        raise ApiError(
            f"Faltan columnas requeridas en CONVENIO: {', '.join(missing)}.",
            status_code=400,
        )

    rows: List[Dict[str, Any]] = []
    errors: List[ConvenioRowError] = []
    total = 0

    for r in range(header_row_idx + 1, (ws.max_row or header_row_idx) + 1):
        raw: Dict[str, Any] = {}
        for col_idx, field in header_map.items():
            raw[field] = ws.cell(row=r, column=col_idx + 1).value

        # fila vacía
        if all(_to_str(raw.get(f)) == "" for f in REQUIRED_FIELDS):
            continue

        total += 1
        out: Dict[str, Any] = {}

        codigo = _parse_codigo(raw.get("destination_codigo"), pad_left=codigo_pad_left)
        kms = _parse_int_strict(raw.get("kms"))
        td = _to_upper(raw.get("td"))

        if not codigo:
            errors.append(
                ConvenioRowError(r, "CLAVE/CÓDIGO inválido (debe ser entero o string sin decimales).", raw)
            )
            continue
        if kms is None:
            errors.append(
                ConvenioRowError(r, "KMS inválido (debe ser entero, sin decimales).", raw)
            )
            continue
        if kms < 0:
            errors.append(ConvenioRowError(r, "KMS debe ser >= 0.", raw))
            continue
        if not td:
            errors.append(ConvenioRowError(r, "TIPO DEST/TD vacío.", raw))
            continue

        out["destination_codigo"] = codigo
        out["kms"] = int(kms)
        out["td"] = td

        # opcionales
        if "destinatario_nombre" in raw:
            name = _to_str(raw.get("destinatario_nombre"))
            out["destinatario_nombre"] = name or None
        if "ciudad" in raw:
            city = _to_str(raw.get("ciudad"))
            out["ciudad"] = city or None

        rows.append(out)

    payload = {
        "ok": len(errors) == 0,
        "sheet": sheet_name,
        "rows": rows,
        "errors": [
            {"row_number": e.row_number, "message": e.message, "data": e.data}
            for e in errors
        ],
        "counts": {"total": total, "valid": len(rows), "invalid": len(errors)},
        "settings": {"codigo_pad_left": int(codigo_pad_left)},
    }
    return payload